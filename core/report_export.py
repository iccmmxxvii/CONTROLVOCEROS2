from __future__ import annotations

from io import BytesIO
from typing import Dict, Mapping

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

GUINDA = "7A1732"
GUINDA_DARK = "551126"
WHITE = "FFFFFF"
SOFT = "F4E8EC"
GRAY = "6B7280"


def dataframe_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def _safe_sheet_name(name: str) -> str:
    clean = str(name).replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "(").replace("]", ")").replace(":", "-")
    return clean[:31] or "Reporte"


def executive_excel_bytes(sheets: Mapping[str, pd.DataFrame], title: str = "ICC Control Territorial") -> bytes:
    """Genera un Excel ejecutivo multihoja a partir del filtro activo."""
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        used = set()
        for raw_name, raw_df in sheets.items():
            name = _safe_sheet_name(raw_name)
            base = name
            idx = 2
            while name in used:
                suffix = f"_{idx}"
                name = (base[: 31 - len(suffix)] + suffix)
                idx += 1
            used.add(name)
            df = raw_df.copy() if isinstance(raw_df, pd.DataFrame) else pd.DataFrame()
            if df.empty:
                df = pd.DataFrame({"Información": ["Sin registros para los filtros seleccionados"]})
            df.to_excel(writer, index=False, sheet_name=name)
            ws = writer.book[name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            ws.sheet_view.showGridLines = False
            ws.row_dimensions[1].height = 28
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=GUINDA)
                cell.font = Font(color=WHITE, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(value), 60))
                    if cell.row > 1:
                        cell.alignment = Alignment(vertical="top", wrap_text=False)
                ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 42))
        writer.book.properties.title = title
        writer.book.properties.subject = "Reporte ejecutivo ICC Control Territorial"
        writer.book.properties.creator = "ICC Control Territorial"
    bio.seek(0)
    return bio.getvalue()
